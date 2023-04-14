
import openpyxl
import random
import matplotlib.pyplot as plt

# 预先设置的参数======这些数都可以按照注释自定义===================
# 车辆发车单价,单位:元
price = 50
# 时间成本，用来权衡最小成本和最高时效的量级，将时间追求转化为成本追求，单位是：元/小时
tprice = 200
# 权重，目标模型是：(1-w) * 时间最短 + w * 成本最低
w = 0.75
# 车辆装载能力（kg）
can = 800
# 单位千克(need_a)、(need_b)、(need_c)每小时对应所耗费的电费
x1 = 0.01
x2 = 0.03
x3 = 0.06
# 每米所耗油费
oil = 0
# 服务单个点的时间,单位：秒
serve_time = 360
# 驾驶员上班时间，如果上班时间没有固定，即满足最早到达第一个点的最早时间窗，那就设置为None
work_start = None
# 迭代次数
times = 500
# 存储配送点信息的类
class Place:
    def __init__(self):
        self.id = 0
        self.name = ''
        self.location = ''
        self.address =  ''
        self.time_a = 0
        self.time_b = 0
        # 货物1
        self.need_a = 0
        # 货物2
        self.need_b = 0
        # 货物3
        self.need_c = 0
        # 被服务完成以及车辆离开的时间
        self.time_served = 0

    # 自身装载总重量
    def load(self):
        sm = self.need_a + self.need_b + self.need_c
        return sm


# 存储配送点信息的类
class Place:
    def __init__(self):
        self.id = 0
        self.name = ''
        self.location = ''
        self.address = ''
        self.time_a = 0
        self.time_b = 0
        # 货物1
        self.need_a = 0
        # 货物2
        self.need_b = 0
        # 货物3
        self.need_c = 0
        # 被服务完成以及车辆离开的时间
        self.time_served = 0

    # 自身装载总重量
    def load(self):
        sm = self.need_a + self.need_b + self.need_c
        return sm


# 存储车辆信息的点
class Car:
    def __init__(self):
        self.id = 0
        # 负责运输的点的队列
        self.array_places = []

    # 返回路线上配送点的数量
    def get_places_len(self):
        return len(self.array_places)

    # 下一次服务的起始时间
    def time_next_start(self):
        return self.array_places[self.get_places_len() - 1].time_served

    # 返回货物1装载重量
    def get_load_need_a(self):
        load = 0
        for index in range(self.get_places_len()):
            load += self.array_places[index].need_a
        return load

    # 返回货物2装载重量
    def get_load_need_b(self):
        load = 0
        for index in range(self.get_places_len()):
            load += self.array_places[index].need_b
        return load

    # 返回货物3装载重量
    def get_load_need_c(self):
        load = 0
        for index in range(self.get_places_len()):
            load += self.array_places[index].need_c
        return load

    # 根据行驶过了几个点，获取剩下的need_a货物的重量，n=0代表还没有服务任何一个点，n=2代表已经服务过两个点，此时车辆上的货物应减少
    def get_load_need_a_by_time(self, n):
        load = 0
        for index in range(n, self.get_places_len()):
            load += self.array_places[index].need_a
        return load

    # 根据行驶过了几个点，获取剩下的need_b货物的重量
    def get_load_need_b_by_time(self, n):
        load = 0
        for index in range(n, self.get_places_len()):
            load += self.array_places[index].need_b
        return load

    # 根据行驶过了几个点，获取剩下的need_c货物的重量
    def get_load_need_c_by_time(self, n):
        load = 0
        for index in range(n, self.get_places_len()):
            load += self.array_places[index].need_c
        return load

    # 返回所有货物重量
    def get_load_all(self):
        return self.get_load_need_a() + self.get_load_need_b() + self.get_load_need_c()

    # 输出该车会走过的里程(不包括起始点出发和回起始点的距离),单位：米
    def get_mileage_places(self):
        if self.get_places_len() == 0:
            return 0
        else:
            arr = self.array_places
            temp = 0
            for index in range(self.get_places_len() - 1):
                temp += get_mileage(arr[index], arr[index + 1])
            return temp

    # 输出该车路径上点的信息
    def output(self):
        for index in self.array_places:
            print("id=", index.id, " name:", index.name, " time_a:", index.time_a, " time_b:", index.time_b,
                  " need_a:", index.need_a, " need_b:", index.need_b, " need_c:", index.need_c, " time_served:",
                  index.time_served)

    # 获得尾部place
    def get_last_place(self):
        return self.array_places[self.get_places_len() - 1]


# 存储起送点信息的类
class Origin:
    def __init__(self):
        self.id = 0
        self.name = ''
        self.location = ''
        self.address = ''
        self.arr_cars = []
        self.paint = []

    # 检测自己的车队是否合格，无超载、满足时间窗,合格的就更新相关信息,如果函数返回True，那么整条链都更新成功
    def is_legal(self):
        flag = True
        for car in self.arr_cars:
            if len(car.array_places) == 0:
                continue
            else:
                if car.get_load_all() > can:
                    flag = False
                    # print(self.id, "号出发点", "car", car.id, "超载，目前", car.get_load_all(), "kg")
                    return flag
                else:
                    arr = car.array_places
                    length = car.get_places_len()
                    if not time_is_legal(arr[0], self):
                        flag = False
                        return flag
                    else:
                        get_new_time_served(arr[0], self)
                        for index in range(length - 1):
                            if not time_is_legal(arr[index + 1], arr[index]):
                                flag = False
                                # print("car", car.id, "时间窗要求不符合")
                                return flag
                            else:
                                get_new_time_served(arr[index + 1], arr[index])
        return flag

    # 输出指定车辆的路径信息
    def output_byid(self, iid):
        if iid > len(self.arr_cars):
            print("没有该车辆，id大了")
        else:
            for index in self.arr_cars:
                if index.id == iid:
                    index.output()
                    break

    # 输出所有车辆的路径信息
    def output_all(self):
        print('\n\n', self.id, "号起点======================================================")
        for index in range(self.get_car_number()):
            print(index + 1, "号车>>>>>>>>>>>>>>>>>>>载重:", self.arr_cars[index].get_load_all(), "kg 里程:",
                  self.get_mileage_byid(index + 1), "m")
            self.arr_cars[index].output()

    # 返回车辆数量
    def get_car_number(self):
        return len(self.arr_cars)

    # 返回指定车辆的载重
    def get_load_byid(self, iid):
        for car in self.arr_cars:
            if car.id == iid:
                return car.get_load_all()
        return None

    # 通过id得到车辆行程距离，包括起始点出发和回起始点的距离
    def get_mileage_byid(self, iid):
        if iid > self.get_car_number():
            print("没有该车辆,id大了")
            return None
        else:
            car = self.arr_cars[iid - 1]
            temp = car.get_mileage_places() + get_mileage(car.array_places[0], self) + get_mileage(self,
                                                                                                   car.get_last_place())
            return temp

    # 获取该起点出发的车辆的所有里程
    def get_mileage_all(self):
        sum_mileage = 0
        for index in range(1, self.get_car_number() + 1):
            sum_mileage += self.get_mileage_byid(index)
        return sum_mileage

    # 获取该起点出发的车辆的电费消耗
    def get_electricity(self):
        # 三种快递总重量×其行驶时间
        sum_ele_1 = 0
        sum_ele_2 = 0
        sum_ele_3 = 0
        for index in self.arr_cars:
            sum_ele_1 += get_time(self, index.array_places[0]) * index.get_load_need_a_by_time(0)
            sum_ele_2 += get_time(self, index.array_places[0]) * index.get_load_need_b_by_time(0)
            sum_ele_3 += get_time(self, index.array_places[0]) * index.get_load_need_c_by_time(0)
            for jndex in range(1, index.get_places_len()):
                sum_ele_1 += (index.array_places[jndex].time_served - index.array_places[
                    jndex - 1].time_served) * index.get_load_need_a_by_time(jndex)
                sum_ele_2 += (index.array_places[jndex].time_served - index.array_places[
                    jndex - 1].time_served) * index.get_load_need_b_by_time(jndex)
                sum_ele_3 += (index.array_places[jndex].time_served - index.array_places[
                    jndex - 1].time_served) * index.get_load_need_c_by_time(jndex)
        sum_ele = sum_ele_1 * x1 + sum_ele_2 * x2 + sum_ele_3 * x3
        return sum_ele

    def get_time_spend(self):
        time_spend = 0
        for index in self.arr_cars:
            time_spend += get_time(self, index.array_places[0])
            for jndex in range(index.get_places_len() - 1):
                time_spend += get_time(index.array_places[jndex], index.array_places[jndex + 1])
            time_spend += get_time(index.get_last_place(), self)
        return time_spend


# 该类用来保存种群和其得分
class Population:
    def __init__(self):
        self.id = 0
        self.arr = []
        self.grades = 0


# 类定义结束============================================================

# 获取两点之间的距离
def get_mileage(place1, place2):
    # place1是origin类
    if is_contain_bylocation(arr_origin_info, place1):
        return float(arr_relation_oTOp[place1.id - 1][place2.id - 1].split(',')[0])
    elif is_contain_bylocation(arr_origin_info, place2):
        return float(arr_relation_oTOp[place2.id - 1][place1.id - 1].split(',')[0])
    # 两个都是place类
    else:
        if place1.id < place2.id:
            return float(arr_relation_pTOp[place1.id - 1][place2.id - 1].split(',')[0])
        else:
            return float(arr_relation_pTOp[place2.id - 1][place1.id - 1].split(',')[0])


# 获取两点之间的时间
def get_time(place1, place2):
    if is_contain_bylocation(arr_origin_info, place1):
        return float(arr_relation_oTOp[place1.id - 1][place2.id - 1].split(',')[1]) / 3600
    elif is_contain_bylocation(arr_origin_info, place2):
        return float(arr_relation_oTOp[place2.id - 1][place1.id - 1].split(',')[1]) / 3600
    else:
        if place1.id < place2.id:
            return float(arr_relation_pTOp[place1.id - 1][place2.id - 1].split(',')[1]) / 3600
        else:
            return float(arr_relation_pTOp[place2.id - 1][place1.id - 1].split(',')[1]) / 3600


# 判断place1能否加到place2点后面，时间是否合法
def time_is_legal(place1, place2):
    # 没有固定工作起始时间
    if work_start is None:
        # place2是origin类
        if is_contain_bylocation(arr_origin_info, place2):
            return True
        # 两个点都是place类
        else:
            if place2.time_served + get_time(place1, place2) + serve_time / 60 / 60 <= place1.time_b:
                return True
            else:
                return False
    # 有固定工作起始时间
    else:
        if is_contain_bylocation(arr_origin_info, place2):
            if work_start + get_time(place1, place2) + serve_time / 60 / 60 <= place1.time_b:
                return True
            else:
                return False
        else:
            if place2.time_served + get_time(place1, place2) + serve_time / 60 / 60 <= place1.time_b:
                return True
            else:
                return False


# 根据一个点，获取后面一个点的time_served
def get_new_time_served(place_new, place_old):
    # 老点是origin
    if is_contain_bylocation(arr_origin_info, place_old):
        # 工作时间不固定
        if work_start is None:
            place_new.time_served = place_new.time_a + serve_time / 60 / 60
        else:
            temp = work_start + get_time(place_old, place_new)
            if temp <= place_new.time_a:
                place_new.time_served = place_new.time_a + serve_time / 60 / 60
            else:
                place_new.time_served = temp + serve_time / 60 / 60
    # 老点是place
    else:
        temp = place_old.time_served + get_time(place_old, place_new)
        if temp <= place_new.time_a:
            place_new.time_served = place_new.time_a + serve_time / 60 / 60
        else:
            place_new.time_served = temp + serve_time / 60 / 60

# 将place加入origin的运输计划中,并且更新place的time_served，删除已经添加的place，该方法也负责判断能否加入
def add_to_origin(place: Place, origin: Origin):
    # 如果origin还没有一辆有计划的车，那么就判断，如果可以运送，就新增一辆车，将place加入该车的运输计划中
    if origin.get_car_number() == 0:
        # 如果时间合法，重量合法(当place作为第一个配送点，重量一定合法，因为一个配送点的需求默认不超过车辆载重)
        if time_is_legal(place, origin):
            car = Car()
            car.id = 1
            get_new_time_served(place, origin)
            car.array_places.append(place)
            origin.arr_cars.append(car)
            arr_place_info.remove(place)
            return True
        else:
            print("place", place.id, "加入失败")
            return False
    # 已经有运输车
    else:
        # 试图加入车辆计划中
        for car in origin.arr_cars:
            if time_is_legal(place, car.get_last_place()) and car.get_load_all() + place.load() <= can:
                get_new_time_served(place, car.get_last_place())
                car.array_places.append(place)
                arr_place_info.remove(place)
                return True
        # 加入车辆计划失败,试图作为第一个配送点，增加配送车辆
        if is_contain_bylocation(arr_place_info, place):
            if time_is_legal(place, origin):
                car = Car()
                car.id = origin.get_car_number() + 1
                get_new_time_served(place, origin)
                car.array_places.append(place)
                origin.arr_cars.append(car)
                arr_place_info.remove(place)
                return True
            else:
                return False


# 输出所有出发点的车辆信息
def output_cars_info(arr):
    time_spend = 0
    road = 0
    for index in arr:
        index.output_all()
        time_spend += index.get_time_spend()
        road += index.get_mileage_all()
    print("总花费时间：", time_spend, "小时，总里程：", road, "米")


# 形成初始可行解,将place加入离他最近的origin中
def init():
    index = 0
    while len(arr_place_info) != 0:
        temp = arr_place_info[0]
        # 逐个寻找该点最近的origin
        min_ori = 0
        min_mileage = 99999999
        for jndex in range(len(arr_origin_info)):
            if get_mileage(temp, arr_origin_info[jndex]) < min_mileage:
                min_mileage = get_mileage(temp, arr_origin_info[jndex])
                min_ori = jndex
        if add_to_origin(temp, arr_origin_info[min_ori]):
            index += 1
            # print(index, "号点添加到", min_ori + 1, "点中")
        else:
            print(index, "号点添加失败")
    if len(arr_place_info) == 0:
        print("全部添加完成")


# 交叉、遗传等
def gene(arr):
    # 随机一个行为：
    #   1代表路径上的某段(或某个)place位置改变
    #   2代表路径上的某段(或某个)place移动到别的路径上，可能是一条新的路径，也可能是已经存在的一条路径
    #   3代表两个路径的某段(或某个)place交换
    ran_do = random.randint(1, 3)
    if ran_do == 1:
        # print("做了do1")
        gene_do1(arr)
    elif ran_do == 2:
        # print("做了do2")
        gene_do2(arr)
    elif ran_do == 3:
        # print("做了do3")
        gene_do3(arr)
    return True


# ran_do == 1时的操作
def gene_do1(arr):
    ran_ori = arr[random.randint(0, len(arr) - 1)]
    # 如果随机到的起点没有车辆计划
    if ran_ori.get_car_number() == 0:
        return True
    else:
        ran_car = ran_ori.arr_cars[random.randint(0, ran_ori.get_car_number() - 1)]
        # 如果车辆的places数小于等于1
        if ran_car.get_places_len() <= 1:
            return True
        # 如果车辆的places大于1
        else:
            # 随机裁切一段数组
            temp = get_arrays(ran_car.array_places)
            # 获取剩下的places的长度
            leng = ran_car.get_places_len()
            # 如果剩下的places==0，那直接插入
            if leng == 0:
                ran_car.array_places.extend(temp)
                return True
            else:
                # 找到随机的一个位置，将裁切下的数组插入
                ran_from = random.randint(0, leng)
                if ran_from == 0:
                    temp.extend(ran_car.array_places)
                    ran_car.array_places = temp
                    return True
                elif ran_from == leng:
                    ran_car.array_places.extend(temp)
                    return True
                else:
                    t1 = ran_car.array_places[0:ran_from]
                    t2 = ran_car.array_places[ran_from:ran_car.get_places_len()]
                    t1.extend(temp)
                    t3 = t1
                    t3.extend(t2)
                    ran_car.array_places = t3
                    return True


# ran_do == 2时的操作
def gene_do2(arr):
    ran_ori_from = arr[random.randint(0, len(arr) - 1)]
    # 如果随机到的点没有车辆计划
    if ran_ori_from.get_car_number() == 0:
        return True
    else:
        ran_car_from = ran_ori_from.arr_cars[random.randint(0, ran_ori_from.get_car_number() - 1)]
        # 如果车辆的点数小于1
        if ran_car_from.get_places_len() == 0:
            return True
        # 如果车辆计划的点数大于等于1
        else:
            # 随机裁切一段places。
            temp = get_arrays(ran_car_from.array_places)
            # 判断car里面是否还有place，如果没有，删除car
            if ran_car_from.get_places_len() == 0:
                ran_ori_from.arr_cars.remove(ran_car_from)
            # 随机抽取一个点，
            ran_ori_to = arr[random.randint(0, len(arr) - 1)]
            # 如果去往的place没有车辆计划
            if ran_ori_to.get_car_number() == 0:
                car = Car()
                car.id = 1
                car.array_places.extend(temp)
                ran_ori_to.arr_cars.append(car)
                return True
            # 如果已有车辆计划
            else:
                # 随机一个数，判断是创建新的车队还是加入现有的车队
                ran_do = random.randint(1, 2)
                # 1的话创建新的车队
                if ran_do == 1:
                    car = Car()
                    car.id = ran_ori_to.get_car_number() + 1
                    car.array_places.extend(temp)
                    ran_ori_to.arr_cars.append(car)
                    return True
                # 2的话加入现有车队
                else:
                    ran_car_to = ran_ori_to.arr_cars[random.randint(0, ran_ori_to.get_car_number() - 1)]
                    # 找到随机的一个位置，将裁切下的数组插入
                    leng = ran_car_to.get_places_len()
                    ran_from = random.randint(0, leng)
                    if ran_from == 0:
                        temp.extend(ran_car_to.array_places)
                        ran_car_to.array_places = temp
                        return True
                    elif ran_from == leng:
                        ran_car_to.array_places.extend(temp)
                        return True
                    else:
                        t1 = ran_car_to.array_places[0:ran_from]
                        t2 = ran_car_to.array_places[ran_from:ran_car_to.get_places_len()]
                        t1.extend(temp)
                        t3 = t1
                        t3.extend(t2)
                        ran_car_to.array_places = t3
                        return True


# ran_do == 3时的操作
def gene_do3(arr):
    # 找到两个不一样的数字，以此找到两个不同的origin
    random_1 = random.randint(0, len(arr) - 1)
    random_2 = random.randint(0, len(arr) - 1)
    while random_2 == random_1:
        random_2 = random.randint(0, len(arr) - 1)
    ran_ori_from = arr[random_1]
    ran_ori_to = arr[random_2]
    # 如果点的车辆数为0，就不作为
    if ran_ori_from.get_car_number() == 0:
        return True
    if ran_ori_to.get_car_number() == 0:
        return True
    else:
        ran_car_from = ran_ori_from.arr_cars[random.randint(0, ran_ori_from.get_car_number() - 1)]
        ran_car_to = ran_ori_to.arr_cars[random.randint(0, ran_ori_to.get_car_number() - 1)]
        temp_from = get_arrays(ran_car_from.array_places)
        temp_to = get_arrays(ran_car_to.array_places)
        # 先处理from到to
        ran_from = random.randint(0, ran_car_to.get_places_len())
        if ran_from == 0:
            temp_from.extend(ran_car_to.array_places)
            ran_car_to.array_places = temp_from
        elif ran_from == ran_car_to.get_places_len():
            ran_car_to.array_places.extend(temp_from)
        else:
            t1 = ran_car_to.array_places[0:ran_from]
            t2 = ran_car_to.array_places[ran_from:ran_car_to.get_places_len()]
            t1.extend(temp_from)
            t3 = t1
            t3.extend(t2)
            ran_car_to.array_places = t3
            # 再处理to到from
        ran_from = random.randint(0, ran_car_to.get_places_len())
        if ran_from == 0:
            temp_to.extend(ran_car_from.array_places)
            ran_car_from.array_places = temp_to
        elif ran_from == ran_car_from.get_places_len():
            ran_car_from.array_places.extend(temp_to)
        else:
            t1 = ran_car_from.array_places[0:ran_from]
            t2 = ran_car_from.array_places[ran_from:ran_car_from.get_places_len()]
            t1.extend(temp_to)
            t3 = t1
            t3.extend(t2)
            ran_car_from.array_places = t3
        return True


# 数组操作，将随机数组的一个或一段提取出来，原数组内容删除
def get_arrays(arr):
    ran_from = random.randint(0, len(arr) - 1)
    ran_to = random.randint(ran_from + 1, len(arr))
    temp = arr[ran_from:ran_to]
    for index in temp:
        arr.remove(index)
    return temp


# 数组操作，复制数组
def copy(arr):
    temp = []
    for index in arr:
        t1 = Origin()
        t1.id = index.id
        t1.name = index.name
        t1.address = index.address
        t1.location = index.location
        t1.paint = index.paint
        for jndex in index.arr_cars:
            t2 = Car()
            t2.id = jndex.id
            for kndex in jndex.array_places:
                t3 = Place()
                t3.id = kndex.id
                t3.name = kndex.name
                t3.address = kndex.address
                t3.location = kndex.location
                t3.time_served = kndex.time_served
                t3.time_a = kndex.time_a
                t3.time_b = kndex.time_b
                t3.need_a = kndex.need_a
                t3.need_b = kndex.need_b
                t3.need_c = kndex.need_c
                t2.array_places.append(t3)
            t1.arr_cars.append(t2)
        temp.append(t1)
    return temp


# 数组操作，判断数组内是否包含该location
def is_contain_bylocation(arr_ori, arr_des):
    for index in arr_ori:
        if index.location == arr_des.location:
            return True
    return False


# 评分函数，用来筛选种群
def get_grades(arr):
    # 里程
    mileage_all =0
    # 车辆数
    car_number = 0
    # 电费
    ele_cost = 0
    # 时间花费,以小时为单位
    time_spend = 0
    for index in arr:
        mileage_all += index.get_mileage_all()
        car_number += index.get_car_number()
        ele_cost += index.get_electricity()
        time_spend += index.get_time_spend()
    price_1 = car_number * price
    price_2 = mileage_all * oil
    price_3 = ele_cost
    return w * (price_1 + price_2 + price_3) + (1 - w) * time_spend * tprice

def new_func():
    return 0


# 合法函数，顾名思义，判断数组内origin的cars是否都合法，不合法全部拉黑
def is_legal(arr):
    for index in arr:
        if not index.is_legal():
            return False
    return True


# 获取路线图
def paint(arr):
    print()
    print()
    print("以下是静态概略图")
    for index in arr:
        if index.get_car_number() != 0:
            print(index.id, "号起点————", index.name)
            for jndex in index.arr_cars:
                string_name = ''
                string = 'https://restapi.amap.com/v3/staticmap?traffic=1&key=17b1bf05ab7fa9683fe2f1a8c822533b&scale=2&location=' + index.location + '&paths=2,0x0000ff,1,,:' + index.location + ';'
                for kndex in jndex.array_places:
                    string += kndex.location + ';'
                    string_name += kndex.name + ' -> '
                string = string[:-1]
                # 以下是添加标签
                # string += "&labels=" + index.name.split('(')[0] + ",1,0,14,0x000FFF,0xFFFFFF:" + index.location + '|'
                # for kndex in jndex.array_places:
                #     string += kndex.name.split('(')[0] + ",1,0,14,0xFFFFFF,0x008000:" + kndex.location + '|'
                # string = string[:-1]
                string += "&labels=" + str(0) + ",1,0,14,0x000FFF,0xFFFFFF:" + index.location + '|'
                ttt = 1
                for kndex in jndex.array_places:
                    string += str(ttt) + ",1,0,14,0xFFFFFF,0x008000:" + kndex.location + '|'
                    ttt += 1
                string = string[:-1]
                print(string_name)
                print(string)
    print()
    print()

## ==================================================
# 读取数据
workbook_place = openpyxl.load_workbook('C:\\Users\\zcz\\Desktop\\py\\地址距离.xlsx')
workbook_origin = openpyxl.load_workbook('C:\\Users\\zcz\\Desktop\\py\\配送中心.xlsx')
# 各点的详细信息
ws_place_info = workbook_place.worksheets[0]
ws_origin_info = workbook_origin.worksheets[0]
# 最大行数
rows_place_info = ws_place_info.max_row
rows_origin_info = ws_origin_info.max_row
# 各点之间的距离,时间
ws_place_relation = workbook_place.worksheets[1]
ws_origin_relation = workbook_origin.worksheets[1]
# 存储各点详细信息的数组
arr_place_info = []
for i in range(2, rows_place_info + 1):
    tmp = Place()
    tmp.id = i - 1
    tmp.name = ws_place_info.cell(i, 3).value
    tmp.location = ws_place_info.cell(i, 4).value
    tmp.address = ws_place_info.cell(i, 2).value
    tmp.time_a = ws_place_info.cell(i, 5).value
    tmp.time_b = ws_place_info.cell(i, 6).value
    tmp.need_a = ws_place_info.cell(i, 7).value
    tmp.need_b = ws_place_info.cell(i, 8).value
    tmp.need_c = ws_place_info.cell(i, 9).value
    arr_place_info.append(tmp)
# 存放起始点信息的数组
arr_origin_info = []
for i in range(1, rows_origin_info + 1):
    tmp = Origin()
    tmp.id = i
    tmp.name = ws_origin_info.cell(i, 3).value
    tmp.location = ws_origin_info.cell(i, 4).value
    tmp.address = ws_origin_info.cell(i, 2).value
    arr_origin_info.append(tmp)
# 存放点到点之间的距离、时间
col_place = ws_place_relation.max_column
col_origin = ws_origin_relation.max_column
rows_place_relation = ws_place_relation.max_row
rows_origin_relation = ws_origin_relation.max_row

arr_relation_pTOp = [[None] * col_place for i in range(rows_place_relation)]
arr_relation_oTOp = [[None] * col_origin for i in range(rows_origin_relation)]
for i in range(rows_place_relation):
    for j in range(col_place):
        arr_relation_pTOp[i][j] = ws_place_relation.cell(i + 1, j + 1).value
for i in range(rows_origin_relation):
    for j in range(col_origin):
        arr_relation_oTOp[i][j] = ws_origin_relation.cell(i + 1, j + 1).value

# 存放路线图的
# 准备工作结束=====================================================================
init()
print("原始分数：", get_grades(arr_origin_info))
output_cars_info(arr_origin_info)
print('开始迭代')
# 运行
# ====================================
# 画目标函数图的坐标存放
pltx = []
plty = []
# ====================================
tt = copy(arr_origin_info)
for i in range(times):
    gene(tt)
    # 因为is_legal()函数也负责更新place的time_served，所以要更新一次，之前这里少了，找bug找了4个多小时
    is_legal(tt)
    com1 = float(get_grades(tt))
    com2 = float(get_grades(arr_origin_info))
    if is_legal(tt) and com1 < com2:
        arr_origin_info = copy(tt)
        pltx.append(i)
        plty.append(com1)
        print(com1)
    else:
        tt = copy(arr_origin_info)
    if i == times - 1:
        print("最终分数：", get_grades(arr_origin_info))

output_cars_info(arr_origin_info)
paint(arr_origin_info)

# 绘目标函数图
plt.plot(pltx, plty, 'c-')
plt.show()

